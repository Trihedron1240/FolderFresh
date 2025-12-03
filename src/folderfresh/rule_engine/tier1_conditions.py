"""
Tier-1 Hazel-style conditions for FolderFresh.

Includes:
- ContentContainsCondition: Match files by text/binary content
- And utilities for date-based pattern matching

All conditions follow FolderFresh patterns:
- Inherit from Condition base class
- Implement evaluate(fileinfo) -> bool
- Print debug info during evaluation
"""

import os
from typing import Dict, Any
from .backbone import Condition


# ============================================================================
# CONTENT-BASED CONDITION
# ============================================================================

class ContentContainsCondition(Condition):
    """Check if file content contains a keyword."""

    def __init__(self, keyword: str, max_bytes: int = 256000):
        """
        Args:
            keyword: Text to search for
            max_bytes: Maximum bytes to read (default 256 KB)
        """
        self.keyword = keyword.lower()
        self.max_bytes = max_bytes

    def evaluate(self, fileinfo: Dict[str, Any]) -> bool:
        """
        Check if file contains the keyword in first max_bytes.

        Supports:
        - Plain text files (.txt, .log, .py, .js, etc.)
        - Binary search for other formats
        - PDF extraction (basic text layer)
        - DOCX extraction (basic text)

        Returns:
            True if keyword found, False otherwise
        """
        file_path = fileinfo.get("path", "")
        file_name = fileinfo.get("name", "")

        if not file_path or not os.path.exists(file_path):
            print(f"  [CONDITION] ContentContains - file not accessible: {file_name}")
            return False

        try:
            ext = os.path.splitext(file_path)[1].lower()

            # Try text extraction based on file type
            content = self._extract_content(file_path, ext)

            if content is None:
                # Fallback: binary search
                try:
                    with open(file_path, 'rb') as f:
                        data = f.read(self.max_bytes)
                    # Search for keyword in binary data
                    keyword_bytes = self.keyword.encode('utf-8', errors='ignore')
                    result = keyword_bytes in data
                except Exception:
                    result = False
            else:
                # Search in extracted text
                result = self.keyword in content.lower()

            print(f"  [CONDITION] ContentContains '{self.keyword}' in {file_name} -> {result}")
            return result

        except Exception as e:
            print(f"  [CONDITION] ContentContains error for {file_name}: {str(e)}")
            return False

    def _extract_content(self, file_path: str, ext: str) -> str | None:
        """
        Extract text content from file based on extension.

        Returns:
            Extracted text (lowercase for comparison), or None if unsupported
        """
        try:
            # Plain text files
            if ext in ['.txt', '.log', '.md', '.rst', '.json', '.xml', '.csv', '.py', '.js', '.java', '.cpp', '.c', '.h', '.cs', '.sh', '.bat', '.ps1']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read(self.max_bytes)

            # PDF extraction (basic - requires pdfplumber or pypdf)
            if ext == '.pdf':
                try:
                    import pdfplumber
                    with pdfplumber.open(file_path) as pdf:
                        text = ""
                        for page in pdf.pages[:5]:  # First 5 pages only
                            text += page.extract_text() or ""
                        return text[:self.max_bytes]
                except ImportError:
                    return None

            # DOCX extraction (basic - requires python-docx)
            if ext == '.docx':
                try:
                    from docx import Document
                    doc = Document(file_path)
                    text = "\n".join(p.text for p in doc.paragraphs)
                    return text[:self.max_bytes]
                except ImportError:
                    return None

            # XLSX extraction (basic - requires openpyxl)
            if ext == '.xlsx':
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    text = ""
                    for sheet in wb.sheetnames[:3]:
                        for row in wb[sheet].iter_rows(values_only=True):
                            text += " ".join(str(c) for c in row if c) + "\n"
                    return text[:self.max_bytes]
                except ImportError:
                    return None

        except Exception:
            pass

        # Unsupported or error
        return None


# ============================================================================
# DATE-BASED PATTERN MATCHING UTILITIES
# ============================================================================

class DatePatternCondition(Condition):
    """
    Check if file's modification/creation date matches a pattern.

    Examples:
    - DatePatternCondition("created", "2024-*")       # Created in 2024
    - DatePatternCondition("modified", "2024-01-*")   # Modified in January 2024
    - DatePatternCondition("created", "*-12-25")      # Created on Christmas
    """

    def __init__(self, date_type: str, pattern: str):
        """
        Args:
            date_type: "created" or "modified"
            pattern: Date pattern like "2024-*", "2024-01-*", "*-12-25"
        """
        self.date_type = date_type.lower()
        self.pattern = pattern

    def evaluate(self, fileinfo: Dict[str, Any]) -> bool:
        """
        Check if file date matches pattern.
        """
        file_path = fileinfo.get("path", "")

        if not file_path or not os.path.exists(file_path):
            return False

        try:
            stat_info = os.stat(file_path)

            # Get timestamp based on date_type
            if self.date_type == "created":
                timestamp = stat_info.st_birthtime if hasattr(stat_info, 'st_birthtime') else stat_info.st_ctime
            elif self.date_type == "modified":
                timestamp = stat_info.st_mtime
            else:
                return False

            from datetime import datetime
            date_str = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d")

            # Match pattern (with wildcard support)
            import fnmatch
            result = fnmatch.fnmatch(date_str, self.pattern)

            print(f"  [CONDITION] DatePattern {self.date_type}='{date_str}' matches '{self.pattern}' -> {result}")
            return result

        except Exception:
            return False
